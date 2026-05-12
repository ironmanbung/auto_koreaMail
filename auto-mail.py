"""
mail_automation.py
메일 자동발송 자동화 프로그램
- 엑셀 로드 및 테이블 편집
- 위치정보 10개 관리 (영속 저장)
- pyautogui 기반 브라우저 자동화
- 화면 좌측 절반 고정 실행
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import threading
import time
import pyautogui
import pyperclip
import openpyxl

# ── 설정 파일 경로 ──────────────────────────────────────────────
CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mail_auto_config.json")

COLUMNS = ["연번", "받는사람이메일", "메일제목", "메일본문", "첨부파일명"]

pyautogui.FAILSAFE = True   # 마우스 좌상단 이동 시 긴급 정지
pyautogui.PAUSE = 0.05

# ── 설정 로드/저장 ──────────────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# ══════════════════════════════════════════════════════════════════
class MailAutomationApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("메일 자동발송 자동화")
        self.root.configure(bg="#1e1e2e")

        # 상태
        self.cfg = load_config()
        self.table_data: list[list] = []       # 엑셀 데이터 (헤더 제외)
        self.attachment_folder = tk.StringVar(value=self.cfg.get("attachment_folder", ""))
        self.stop_flag = threading.Event()
        self.capturing_index: int | None = None  # 현재 캡처 중인 위치정보 인덱스(0-base)

        # 위치정보 저장소 (10개)
        self.pos_vars   = []   # StringVar – "x,y"
        self.memo_vars  = []   # StringVar – 메모
        saved_pos   = self.cfg.get("positions",  [""] * 10)
        saved_memo  = self.cfg.get("memos",      [""] * 10)
        for i in range(10):
            pv = tk.StringVar(value=saved_pos[i]  if i < len(saved_pos)  else "")
            mv = tk.StringVar(value=saved_memo[i] if i < len(saved_memo) else "")
            self.pos_vars.append(pv)
            self.memo_vars.append(mv)

        # 화면 좌측 절반에 창 배치
        self._position_window()

        # UI 구성
        self._build_ui()

        # 종료 시 저장
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── 창 위치 설정 ────────────────────────────────────────────
    def _position_window(self):
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        w  = sw // 2
        h  = sh
        self.root.geometry(f"{w}x{h}+0+0")
        self.root.resizable(False, False)

    # ── UI 빌드 ─────────────────────────────────────────────────
    def _build_ui(self):
        BG   = "#1e1e2e"
        FG   = "#cdd6f4"
        ACC  = "#89b4fa"
        BTN  = "#313244"
        BTNF = "#cdd6f4"
        SEL  = "#45475a"

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background="#181825", foreground=FG,
                        fieldbackground="#181825", rowheight=24,
                        font=("Malgun Gothic", 9))
        style.configure("Treeview.Heading",
                        background=BTN, foreground=ACC,
                        font=("Malgun Gothic", 9, "bold"))
        style.map("Treeview", background=[("selected", SEL)])

        def btn(parent, text, cmd, width=None, bg=BTN, fg=BTNF, **kw):
            b = tk.Button(parent, text=text, command=cmd,
                          bg=bg, fg=fg, relief="flat",
                          font=("Malgun Gothic", 9, "bold"),
                          cursor="hand2", padx=6, pady=4,
                          activebackground=ACC, activeforeground="#1e1e2e",
                          **kw)
            if width:
                b.config(width=width)
            return b

        def label(parent, text, **kw):
            defaults = dict(bg=BG, fg=FG, font=("Malgun Gothic", 9))
            defaults.update(kw)  # 호출 측 값이 기본값을 덮어씀
            return tk.Label(parent, text=text, **defaults)

        def entry(parent, textvariable, width=20, **kw):
            return tk.Entry(parent, textvariable=textvariable,
                            bg="#313244", fg=FG, insertbackground=FG,
                            relief="flat", font=("Malgun Gothic", 9),
                            width=width, **kw)

        # ── 섹션 프레임 헬퍼 ──
        def section(parent, title):
            f = tk.LabelFrame(parent, text=f"  {title}  ",
                              bg=BG, fg=ACC,
                              font=("Malgun Gothic", 9, "bold"),
                              relief="groove", bd=1)
            return f

        main = tk.Frame(self.root, bg=BG)
        main.pack(fill="both", expand=True, padx=8, pady=8)

        # ┌─────────────────────────────────────────┐
        # │ 1. 엑셀 파일 로드                        │
        # └─────────────────────────────────────────┘
        sec1 = section(main, "① 엑셀 파일")
        sec1.pack(fill="x", pady=(0, 6))

        row1 = tk.Frame(sec1, bg=BG)
        row1.pack(fill="x", padx=6, pady=4)

        self.excel_path_var = tk.StringVar(value=self.cfg.get("excel_path", ""))
        entry(row1, self.excel_path_var, width=45).pack(side="left", padx=(0, 6))
        btn(row1, "📂 엑셀 선택", self._load_excel).pack(side="left", padx=(0, 4))
        btn(row1, "💾 테이블 저장", self._save_table_to_excel).pack(side="left")

        # ┌─────────────────────────────────────────┐
        # │ 2. 테이블                                │
        # └─────────────────────────────────────────┘
        sec2 = section(main, "② 데이터 테이블 (더블클릭으로 셀 편집)")
        sec2.pack(fill="both", expand=True, pady=(0, 6))

        tree_frame = tk.Frame(sec2, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=4)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")

        self.tree = ttk.Treeview(tree_frame,
                                 columns=COLUMNS,
                                 show="headings",
                                 yscrollcommand=vsb.set,
                                 xscrollcommand=hsb.set,
                                 selectmode="browse",
                                 height=10)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        col_widths = [40, 160, 130, 180, 120]
        for col, w in zip(COLUMNS, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40, anchor="w")

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._on_tree_double_click)

        # 테이블 조작 버튼
        tbl_btn_row = tk.Frame(sec2, bg=BG)
        tbl_btn_row.pack(fill="x", padx=6, pady=(0, 4))
        btn(tbl_btn_row, "➕ 행 추가", self._add_row).pack(side="left", padx=(0, 4))
        btn(tbl_btn_row, "➖ 선택행 삭제", self._del_row).pack(side="left")

        # ┌─────────────────────────────────────────┐
        # │ 3. 첨부파일 폴더                         │
        # └─────────────────────────────────────────┘
        sec3 = section(main, "③ 첨부파일 폴더")
        sec3.pack(fill="x", pady=(0, 6))

        row3 = tk.Frame(sec3, bg=BG)
        row3.pack(fill="x", padx=6, pady=4)

        entry(row3, self.attachment_folder, width=45).pack(side="left", padx=(0, 6))
        btn(row3, "📁 폴더 선택", self._select_folder).pack(side="left", padx=(0, 4))
        btn(row3, "📋 경로 복사", self._copy_folder_path).pack(side="left")

        # ┌─────────────────────────────────────────┐
        # │ 4. 위치정보 (10개)                       │
        # └─────────────────────────────────────────┘
        sec4 = section(main, "④ 위치정보 설정 (버튼 클릭 후 목표 위치를 마우스로 클릭)")
        sec4.pack(fill="x", pady=(0, 6))

        pos_frame = tk.Frame(sec4, bg=BG)
        pos_frame.pack(fill="x", padx=6, pady=4)

        col_left  = tk.Frame(pos_frame, bg=BG)
        col_right = tk.Frame(pos_frame, bg=BG)
        col_left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        col_right.pack(side="left", fill="both", expand=True)

        self.pos_capture_btns = []
        self.pos_status_labels = []

        for i in range(10):
            parent_col = col_left if i < 5 else col_right
            row = tk.Frame(parent_col, bg=BG)
            row.pack(fill="x", pady=1)

            capture_btn = btn(row, f"위치{i+1} 지정",
                              lambda idx=i: self._start_capture(idx),
                              width=11)
            capture_btn.pack(side="left", padx=(0, 3))
            self.pos_capture_btns.append(capture_btn)

            status = label(row, "●", fg="#6c7086")
            status.pack(side="left", padx=(0, 2))
            self.pos_status_labels.append(status)

            pos_entry = entry(row, self.pos_vars[i], width=10)
            pos_entry.pack(side="left", padx=(0, 4))

            label(row, "메모:").pack(side="left")
            memo_entry = entry(row, self.memo_vars[i], width=12)
            memo_entry.pack(side="left", padx=(0, 2))

        save_pos_btn_row = tk.Frame(sec4, bg=BG)
        save_pos_btn_row.pack(fill="x", padx=6, pady=(2, 4))
        btn(save_pos_btn_row, "💾 위치정보 저장", self._save_positions).pack(side="left")

        # ┌─────────────────────────────────────────┐
        # │ 5. 메일 발송 제어                        │
        # └─────────────────────────────────────────┘
        sec5 = section(main, "⑤ 메일 발송")
        sec5.pack(fill="x", pady=(0, 4))

        mail_row = tk.Frame(sec5, bg=BG)
        mail_row.pack(fill="x", padx=6, pady=6)

        btn(mail_row, "📧 선택행 1건 메일보내기",
            self._send_selected,
            bg="#a6e3a1", fg="#1e1e2e").pack(side="left", padx=(0, 6))

        btn(mail_row, "📨 전부 메일보내기",
            self._send_all,
            bg="#89b4fa", fg="#1e1e2e").pack(side="left", padx=(0, 6))

        self.stop_btn = btn(mail_row, "⛔ 중지",
                            self._stop_sending,
                            bg="#f38ba8", fg="#1e1e2e")
        self.stop_btn.pack(side="left")

        # 상태바
        self.status_var = tk.StringVar(value="준비")
        status_bar = tk.Label(main, textvariable=self.status_var,
                              bg="#11111b", fg="#a6e3a1",
                              font=("Malgun Gothic", 9),
                              anchor="w", padx=6)
        status_bar.pack(fill="x", pady=(4, 0))

        # 초기 엑셀 로드 (저장된 경로 있을 때)
        if self.excel_path_var.get() and os.path.exists(self.excel_path_var.get()):
            self._load_excel_from_path(self.excel_path_var.get())

    # ══════════════════════════════════════════════════════════════
    # ① 엑셀 로드 / 저장
    # ══════════════════════════════════════════════════════════════
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")]
        )
        if not path:
            return
        self.excel_path_var.set(path)
        self._load_excel_from_path(path)

    def _load_excel_from_path(self, path):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messagebox.showwarning("경고", "엑셀 파일이 비어 있습니다.")
                return
            # 첫 행 = 헤더 검증
            header = [str(c) if c else "" for c in rows[0]]
            if header != COLUMNS:
                messagebox.showwarning(
                    "헤더 불일치",
                    f"헤더가 다릅니다.\n기대: {COLUMNS}\n파일: {header}\n\n계속 로드합니다."
                )
            self.table_data = [[str(c) if c is not None else "" for c in row]
                               for row in rows[1:]]
            self._refresh_tree()
            self._set_status(f"✅ 로드 완료: {len(self.table_data)}건")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 로드 실패:\n{e}")

    def _save_table_to_excel(self):
        path = self.excel_path_var.get()
        if not path:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")]
            )
            if not path:
                return
            self.excel_path_var.set(path)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(COLUMNS)
            for row in self.table_data:
                ws.append(row)
            wb.save(path)
            self._set_status(f"💾 저장 완료: {path}")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패:\n{e}")

    # ── 트리뷰 갱신 ─────────────────────────────────────────────
    def _refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.table_data:
            # 5열 보장
            r = row[:5] + [""] * (5 - len(row))
            self.tree.insert("", "end", values=r)

    # ── 셀 더블클릭 편집 ────────────────────────────────────────
    def _on_tree_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)  # '#1' ~ '#5'
        if not row_id or not col_id:
            return

        col_idx = int(col_id.replace("#", "")) - 1
        row_idx = self.tree.index(row_id)
        current = self.table_data[row_idx][col_idx] if col_idx < len(self.table_data[row_idx]) else ""

        # bbox 위에 Entry 팝업
        bbox = self.tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        edit_var = tk.StringVar(value=current)
        edit_entry = tk.Entry(self.tree, textvariable=edit_var,
                              bg="#45475a", fg="#cdd6f4",
                              insertbackground="#cdd6f4",
                              relief="flat", font=("Malgun Gothic", 9))
        edit_entry.place(x=x, y=y, width=w, height=h)
        edit_entry.focus_set()
        edit_entry.select_range(0, "end")

        def commit(e=None):
            val = edit_var.get()
            while len(self.table_data[row_idx]) <= col_idx:
                self.table_data[row_idx].append("")
            self.table_data[row_idx][col_idx] = val
            self._refresh_tree()
            edit_entry.destroy()

        def cancel(e=None):
            edit_entry.destroy()

        edit_entry.bind("<Return>", commit)
        edit_entry.bind("<Tab>", commit)
        edit_entry.bind("<Escape>", cancel)
        edit_entry.bind("<FocusOut>", commit)

    # ── 행 추가/삭제 ────────────────────────────────────────────
    def _add_row(self):
        n = len(self.table_data) + 1
        self.table_data.append([str(n), "", "", "", ""])
        self._refresh_tree()

    def _del_row(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("알림", "삭제할 행을 선택하세요.")
            return
        idx = self.tree.index(sel[0])
        del self.table_data[idx]
        self._refresh_tree()

    # ══════════════════════════════════════════════════════════════
    # ③ 첨부파일 폴더
    # ══════════════════════════════════════════════════════════════
    def _select_folder(self):
        folder = filedialog.askdirectory(title="첨부파일 폴더 선택")
        if folder:
            self.attachment_folder.set(folder)
            self._set_status(f"📁 폴더 지정: {folder}")

    def _copy_folder_path(self):
        path = self.attachment_folder.get()
        if path:
            pyperclip.copy(path)
            self._set_status(f"📋 클립보드 복사: {path}")
        else:
            messagebox.showinfo("알림", "폴더를 먼저 선택하세요.")

    # ══════════════════════════════════════════════════════════════
    # ④ 위치정보 캡처
    # ══════════════════════════════════════════════════════════════
    def _start_capture(self, idx: int):
        """
        버튼 클릭 → 1.5초 대기 후 마우스 위치 폴링 시작.
        실제 클릭은 pyautogui 마우스 상태 감지로 판별.
        pynput을 쓰지 않아 Tkinter 충돌 없음.
        """
        if self.capturing_index is not None:
            return

        self.capturing_index = idx
        self.pos_capture_btns[idx].config(bg="#fab387", fg="#1e1e2e")
        self.pos_status_labels[idx].config(fg="#fab387")

        # 카운트다운: 버튼에서 손가락을 떼고 이동할 시간 1.5초 확보
        self._set_status(f"🎯 위치{idx+1} — 1.5초 후 클릭 감지 시작. 지금 목표 위치로 이동하세요!")
        self._capture_countdown = 30  # 30초 타임아웃 (×100ms = 3000회)
        self._prev_mouse_state  = True   # 버튼을 누른 채 시작하므로 초기값=True
        self.root.after(1500, lambda: self._poll_mouse_click(idx))

    def _poll_mouse_click(self, idx: int):
        """
        100ms 간격으로 마우스 왼쪽 버튼 상태를 폴링.
        눌림(True) → 안눌림(False) → 다시 눌림(True) 전환을 감지 = 새로운 클릭.
        """
        if self.capturing_index != idx:
            return  # 취소됨

        try:
            import ctypes
            # GetAsyncKeyState: 마우스 왼쪽 버튼(VK=0x01) 상태
            state = bool(ctypes.windll.user32.GetAsyncKeyState(0x01) & 0x8000)
        except Exception:
            # fallback: pyautogui position only (Windows 외 OS)
            state = False

        prev = getattr(self, "_prev_mouse_state", False)

        if not prev and state:
            # 버튼이 눌린 순간 → 현재 마우스 좌표 저장
            x, y = pyautogui.position()
            self.root.after(0, lambda: self._finish_capture(idx, (x, y)))
            return

        self._prev_mouse_state = state
        self._capture_countdown -= 1

        if self._capture_countdown <= 0:
            self.root.after(0, lambda: self._finish_capture(idx, None))
            return

        self.root.after(100, lambda: self._poll_mouse_click(idx))

    def _finish_capture(self, idx: int, pos):
        self.capturing_index = None
        # 버튼 색 복원
        self.pos_capture_btns[idx].config(bg="#313244", fg="#cdd6f4")

        if pos:
            self.pos_vars[idx].set(f"{pos[0]},{pos[1]}")
            self.pos_status_labels[idx].config(fg="#a6e3a1")
            self._set_status(f"✅ 위치정보{idx+1} 저장: {pos[0]},{pos[1]}")
            self._save_positions()
        else:
            self.pos_status_labels[idx].config(fg="#f38ba8")
            self._set_status(f"⏰ 위치정보{idx+1} 캡처 타임아웃")

    def _save_positions(self):
        self.cfg["positions"] = [v.get() for v in self.pos_vars]
        self.cfg["memos"]     = [v.get() for v in self.memo_vars]
        save_config(self.cfg)
        self._set_status("💾 위치정보 저장 완료")

    def _get_pos(self, idx: int):
        """위치정보 idx(0-base) → (x, y) 또는 None"""
        val = self.pos_vars[idx].get().strip()
        if not val:
            return None
        try:
            x, y = val.split(",")
            return int(float(x)), int(float(y))
        except Exception:
            return None

    # ══════════════════════════════════════════════════════════════
    # ⑤ 메일 발송
    # ══════════════════════════════════════════════════════════════
    def _send_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("알림", "발송할 행을 선택하세요.")
            return
        idx = self.tree.index(sel[0])
        self.stop_flag.clear()
        t = threading.Thread(target=self._send_rows, args=([idx],), daemon=True)
        t.start()

    def _send_all(self):
        if not self.table_data:
            messagebox.showinfo("알림", "데이터가 없습니다.")
            return
        self.stop_flag.clear()
        indices = list(range(len(self.table_data)))
        t = threading.Thread(target=self._send_rows, args=(indices,), daemon=True)
        t.start()

    def _stop_sending(self):
        self.stop_flag.set()
        self._set_status("⛔ 중지 요청됨")

    # ══════════════════════════════════════════════════════════════
    # ⑥ 자동화 핵심 로직
    # ══════════════════════════════════════════════════════════════
    def _send_rows(self, indices: list[int]):
        """
        발송 스레드 – 조건6 시퀀스 수행
        위치정보 인덱스 (0-base):
          0 → 위치정보1  : 첫 버튼 클릭 (메일 작성 시작 버튼 등)
          1 → 위치정보2  : 받는사람 입력 필드
          2 → 위치정보3  : 메일제목 입력 필드
          3 → 위치정보4  : 메일본문 입력 필드
          4 → 위치정보5  : (end키 후 다음으로)
          5 → 위치정보6  : 첨부파일 버튼
          6 → 위치정보7  : 파일 경로 입력창
          7 → 위치정보8  : 파일명 입력 필드
          8 → 위치정보9  : 발송 버튼
        """
        folder = self.attachment_folder.get()

        for n, row_idx in enumerate(indices):
            if self.stop_flag.is_set():
                self.root.after(0, lambda: self._set_status("⛔ 발송 중지됨"))
                return

            row = self.table_data[row_idx]
            # 5열 보장
            while len(row) < 5:
                row.append("")
            _, email, subject, body, filename = row[:5]

            self.root.after(0, lambda i=row_idx, tot=len(indices), cur=n+1:
                            self._set_status(f"📧 발송 중 ({cur}/{tot}) - {self.table_data[i][1]}"))

            try:
                self._automate_mail(email, subject, body, filename, folder)
            except Exception as e:
                self.root.after(0, lambda err=e:
                                messagebox.showerror("자동화 오류", str(err)))
                return

            # 다음 레코드 전 대기 (0.5초)
            if n < len(indices) - 1:
                time.sleep(0.5)

        self.root.after(0, lambda: self._set_status(f"✅ {len(indices)}건 발송 완료"))

    def _automate_mail(self, email, subject, body, filename, folder):
        """
        pyautogui 자동화 시퀀스 (조건6)
        ※ 브라우저는 화면 우측 절반에 미리 배치되어 있어야 합니다.
        """
        DELAY = 0.5

        def click_pos(idx, desc=""):
            pos = self._get_pos(idx)
            if pos is None:
                raise ValueError(f"위치정보{idx+1}이(가) 설정되지 않았습니다. ({desc})")
            pyautogui.click(pos[0], pos[1])

        def paste_text(text):
            pyperclip.copy(str(text))
            time.sleep(0.1)
            pyautogui.hotkey("ctrl", "v")

        # 1) 위치정보1 클릭 (메일 작성 버튼 등)
        click_pos(0, "메일 작성 시작")
        time.sleep(DELAY)
        pyautogui.press("enter")
        time.sleep(DELAY)

        # 2) 위치정보2 클릭 → 받는사람 이메일 붙여넣기
        click_pos(1, "받는사람")
        time.sleep(DELAY)
        paste_text(email)
        time.sleep(DELAY)

        # 3) 위치정보3 클릭 → 메일제목 붙여넣기
        click_pos(2, "메일제목")
        time.sleep(DELAY)
        paste_text(subject)
        time.sleep(DELAY)

        # 4) 위치정보4 클릭 → 메일본문 붙여넣기
        click_pos(3, "메일본문")
        time.sleep(DELAY)
        paste_text(body)
        time.sleep(DELAY)

        # 5) 위치정보5 클릭 → End 키
        click_pos(4, "본문 끝")
        time.sleep(DELAY)
        pyautogui.press("end")
        time.sleep(DELAY)

        # 6) 위치정보6 클릭 (첨부파일 버튼)
        click_pos(5, "첨부파일 버튼")
        time.sleep(DELAY)

        # 7) 위치정보7 클릭 → 폴더 경로 붙여넣기
        click_pos(6, "파일경로 입력창")
        time.sleep(DELAY)
        paste_text(folder)
        time.sleep(DELAY)

        # 8) 위치정보8 클릭 → 파일명 붙여넣기 → Enter → Esc
        click_pos(7, "파일명 입력")
        time.sleep(DELAY)
        paste_text(filename)
        pyautogui.press("enter")
        time.sleep(DELAY)
        pyautogui.press("escape")
        time.sleep(DELAY)

        # 9) 위치정보9 클릭 (발송 버튼)
        click_pos(8, "발송 버튼")

    # ══════════════════════════════════════════════════════════════
    # 유틸
    # ══════════════════════════════════════════════════════════════
    def _set_status(self, msg: str):
        self.status_var.set(msg)

    def _on_close(self):
        """종료 시 모든 설정 저장"""
        self.cfg["excel_path"]         = self.excel_path_var.get()
        self.cfg["attachment_folder"]  = self.attachment_folder.get()
        self.cfg["positions"]          = [v.get() for v in self.pos_vars]
        self.cfg["memos"]              = [v.get() for v in self.memo_vars]
        save_config(self.cfg)
        self.root.destroy()


# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app = MailAutomationApp(root)
    root.mainloop()
